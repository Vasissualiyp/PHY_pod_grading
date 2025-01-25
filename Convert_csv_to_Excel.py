import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

def convert_csv_to_excel(csv_path, sample_excel_path, output_excel_path, column_mapping=None):
    """
    Convert a CSV file to an Excel format based on a sample Excel file.

    Parameters:
    - csv_path: Path to the input CSV file.
    - sample_excel_path: Path to the sample Excel file that provides the format.
    - output_excel_path: Path to save the converted Excel file.
    - column_mapping: Optional dictionary to map CSV columns to Excel columns.
    """
    
    def apply_color_scheme(ws, color_scheme):
        # Apply header colors
        for col_idx, column in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
            for cell in column:
                if cell.value in color_scheme and isinstance(color_scheme[cell.value], str) and color_scheme[cell.value] != '00000000':
                    cell.fill = PatternFill(start_color=color_scheme[cell.value], end_color=color_scheme[cell.value], fill_type="solid")

        # Assume alternating row colors based on the first few rows in the sample
        previous_color = None
        alternating_colors = []

        for row in ws.iter_rows(min_row=2, max_row=5):
            current_color = row[0].fill.start_color.rgb
            if current_color != previous_color:
                alternating_colors.append(current_color)
                previous_color = current_color

        # Apply alternating row colors
        color_idx = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            fill_color = alternating_colors[color_idx]
            for cell in row:
                if fill_color != '00000000':  # 00000000 indicates no fill
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            color_idx = (color_idx + 1) % len(alternating_colors)
    
    # Load the data and trim after the 'Section' column
    csv_data = pd.read_csv(csv_path)
    if 'Section' in csv_data.columns:
        csv_data = csv_data.loc[:, :'Section']
    
    # Rename columns based on mapping
    if column_mapping:
        csv_data = csv_data.rename(columns=column_mapping)
    
    # Load the workbook
    wb = load_workbook(sample_excel_path)
    
    # Select the 'Pods' sheet
    if 'Pods' in wb.sheetnames:
        ws = wb['Pods']
    else:
        ws = wb.create_sheet('Pods')
    
    # Clear previous data
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)  # Remove any previous coloring

    # Append new data
    for r_idx, row in enumerate(dataframe_to_rows(csv_data, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Extract and apply the color scheme
    color_scheme = {}
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value and cell.fill.start_color:
                color_scheme[cell.value] = cell.fill.start_color.rgb
    apply_color_scheme(ws, color_scheme)

    # Remove color for "Original Pod #" and "Pod #" columns
    for col in ['Original Pod #', 'Pod #']:
        if col in color_scheme:
            del color_scheme[col]

    # Apply conditional formatting to "Pod #" columns for different pods
    pod_col_letter = None
    for column in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1, max_row=1):
        for cell in column:
            if cell.value == "Pod #":
                pod_col_letter = cell.column_letter

    if pod_col_letter:
        rule = ColorScaleRule(start_type="min", start_color="E5E5E5", end_type="max", end_color="0000FF")
        ws.conditional_formatting.add(f"{pod_col_letter}2:{pod_col_letter}{ws.max_row}", rule)
    
    # Save to the output file
    wb.save(output_excel_path)
    return output_excel_path

# Example usage of the function
convert_csv_to_excel("./Practical 2/PHY151_Pods_Prac2.csv", "./Sample Pods File.xlsx", "./Converted_Pods_File_Function.xlsx")

