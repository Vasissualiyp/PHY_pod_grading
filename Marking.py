# Libraries 
import numpy as np
import os
import time
import pandas as pd
import configparser
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import csv

from Grading_Schemes import set_grade, set_grade_individual_scheme

def import_Quercus():
    """
    Imports Quercus-specific functions
    """
    from Quercus import import_grades_to_quercus, login_to_quercus, get_assignment_id, get_driver
    try:
        from getpass4 import getpass
    except ImportError:
        from getpass import getpass

# Import configuration 
class PhyConfig():
    '''A class that holds config data'''
    def __init__(self, config_file="config.txt"):
        '''Create a new instance of the class'''
        config = configparser.ConfigParser()
        config.read(config_file)
        
        # Get the values from the PRA section
        PRA_no = config.get('PRA', 'PRA_no')
        
        # Get the values from the QUERCUS section
        self.quercus_grading = config.get('QUERCUS', 'quercus_grading')
        self.login = config.get('QUERCUS', 'login')
        self.password = config.get('QUERCUS', 'password')
        self.browser = config.get('QUERCUS', 'browser')
        self.TFA = config.get('QUERCUS', 'TFA')
        self.webpg_course = config.get('QUERCUS', 'webpg_course')
        self.student_group = config.get('QUERCUS', 'student_group')
        
        # Get the values from the FILES section
        file_in = config.get('FILES', 'file_in')
        file_out = config.get('FILES', 'file_out')
        self.xlsx_colors = config.get('FILES', 'xlsx_colors').split(',')
        
        # Get the values from the GRADING section
        PRA_name_prefix = config.get('GRADING', 'PRA_name_prefix')
        self.GradingScheme = config.get('GRADING', 'GradingScheme')

        self.PRA_name = PRA_name_prefix + PRA_no
        self.file_in  = os.path.join(self.PRA_name, file_in)
        self.file_out = os.path.join(self.PRA_name, file_out)

# Functions definitions 
class Marking():
    '''A class that performs marking and grade upload to Quercus'''
    def __init__(self, config_file="config.txt", removerows=0):
        '''Read config and add variables to the Marking class'''
        self.config = PhyConfig(config_file)
        self.individual_grades_datasets = [ "PHY131_Practical_W2025" ]
        self.removerows = removerows

    # Reading excel file 
    def define_dataframes(self):
    
        # Students list
        dated = pd.read_excel(self.config.file_in, sheet_name="Pods")
        df_students = dated.values.tolist()
        df_students = list(map(list, zip(*df_students)))
    
        # Pod marks
        dated2 = pd.read_excel(self.config.file_in, sheet_name="Marks")
        df_marks = dated2.values.tolist()
        df_marks = list(map(list, zip(*df_marks)))
    
        # Remove rows at the start
        df_students = [row[self.removerows:] for row in df_students]
    
        # Names for the columns in final excel file
        self.Names = list(dated.columns.values)
        if self.config.GradingScheme in self.individual_grades_datasets:
            name1 = self.config.PRA_name + " - Practical questions"
            name2 = self.config.PRA_name + " - notebook"
            name3 = self.config.PRA_name + " - individual"
            name4 = self.config.PRA_name + " - Lateness"
            self.Names.append(name1)
            self.Names.append(name2)
            self.Names.append(name3)
            self.Names.append(name4)
        else:
            self.Names.append(self.config.PRA_name)
    
        return df_students, df_marks
    
    # Grading routine
    def grading(self, df_students, df_marks): 
    
        # Extracting info from excel file 
        # Student List
        #students_last = df_students[lastnamecolumn]
        #students_first = df_students[firstnamecolumn]
        namecolumn = self.Names.index('Student')
        students_name = df_students[namecolumn]
        podnocolumn = self.Names.index('Pod #, 0 if absent')
        students_pod = df_students[podnocolumn]
        latenesscolumn = self.Names.index('Lateness') 
        if self.config.GradingScheme == 'Grades_Column':
            gradescolumn = self.Names.index('Grades') 
            lateness = df_students[gradescolumn]
        else:
            lateness = df_students[latenesscolumn]
        max_pod = max(students_pod)
        # print(students_pod)
    
        # Marks
        Excel_podno = df_marks[0]
        Excel_podmarks = df_marks[1]
        Excel_size = len(Excel_podmarks)
        print(Excel_podmarks)
    
        # Get the marks for each pod 
        # Initialize the array with zeros and max_pod+1 elements
        Pod_marks = [0] * (max_pod + 1)
        if self.config.GradingScheme in self.individual_grades_datasets:
            Excel_podmarks_extra = df_marks[2]
            Pod_marks_extra = [0] * (max_pod + 1)
        # Get the marks for each pod
        for i in range(0, Excel_size):
            # print(i)
            PodNo = int(Excel_podno[i])
            Pod_marks[PodNo] = Excel_podmarks[i]
            if self.config.GradingScheme in self.individual_grades_datasets:
                Pod_marks_extra[PodNo] = Excel_podmarks_extra[i]
    
        # print(Pod_marks)
        
        # Editing the column with marks 
        # Creating a column with marks
        marks = np.zeros(np.size(students_pod))
        df_students.append(marks)
        if self.config.GradingScheme in self.individual_grades_datasets:
            marks_extra = np.zeros_like(marks)
            df_students.append(marks_extra) # For notebook marks
            df_students.append(marks_extra) # For lateness
            df_students.append(marks_extra) # For individual marks
    
        # Do the grading according to the scheme
        if self.config.GradingScheme == 'Grades Column':
            lateness = df_students
        if self.config.GradingScheme not in self.individual_grades_datasets: 
            set_grade(students_pod, Pod_marks, marks, lateness, self.config.GradingScheme)
        else:
            marks_list = set_grade_individual_scheme(students_pod, Pod_marks, Pod_marks_extra, marks, marks_extra, lateness, self.config.GradingScheme)
    
        # Put data into the dataframe 
        if self.config.GradingScheme in self.individual_grades_datasets:
            print(f"0: {marks_list[0]}")
            print(f"1: {marks_list[1]}")
            print(f"2: {marks_list[2]}")
            print(f"3: {marks_list[3]}")
            df_students[9]  = marks_list[0]
            df_students[10] = marks_list[1]
            df_students[11] = marks_list[2]
            df_students[12] = marks_list[3]
        else:
            df_students[9] = marks
        df = pd.DataFrame(df_students)
        df = df.T
        df.columns = self.Names
        df = df.drop(
            [
                "Original Pod #",
                "Pod #, 0 if absent",
                "Lateness"
            ],
            axis=1,
        )
        #NewNames = ["Last", "First", "Pod#", "Late", "Mark"]
        #df.columns = NewNames
        return df
    
    # Apply alternating coloring to the excel file of the dataframe 
    def write_to_excel_with_alternating_colors(self, df, color_list, filename):
        # Create a new workbook
        wb = Workbook()
    
        # Select the active worksheet
        ws = wb.active
    
        # Write the DataFrame to the worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
    
        # Set alternating row colors
        #for i, row in enumerate(ws.iter_rows(min_row=2)):
        #    fill = PatternFill(
        #        start_color=color_list[i % len(color_list)],
        #        end_color=color_list[i % len(color_list)],
        #        fill_type="solid",
        #    )
        #    for cell in row:
        #        cell.fill = fill
    
        # Save the workbook
        wb.save(filename)
        print(f"Saved the output into {filename}!")
    
    #Write the dataframe into the CSV file 
    def write_to_csv(self, df, filename):
        # Open the output file
        with open(filename, "w", newline="", encoding="utf-8") as f:
            # Create a CSV writer object
            writer = csv.writer(f)
    
            # Write the DataFrame headers to the CSV file
            writer.writerow(df.columns)
    
            # Write the DataFrame rows to the CSV file
            for i, row in df.iterrows():
                writer.writerow(row)

    def main(self):
        '''
        Main loop of the marking class, that should do all the marking
        '''
        # Create dataframe with all the marks 
        
        # Import data from excel file into the dataframes
        df_students, df_marks = marking.define_dataframes()
        print(self.Names)
        
        #grading
        df = self.grading(df_students, df_marks)
        
        # Output the dataframe into the file 
        # Get the extension of the file
        file_extension = os.path.splitext(self.config.file_out)[1]
        file_extension = file_extension[1:]
        
        # Excel output
        if file_extension == 'xlsx':
            self.write_to_excel_with_alternating_colors(df, self.config.xlsx_colors, self.config.file_out)
        
        # csv output + Quercus marks import
        elif file_extension == 'csv' and self.config.quercus_grading:
            import_Quercus()
            if self.config.login =='':
                login = input("Enter your login: ")
            if self.config.password == '':
                password = getpass("Enter your password: ")
            auth_info = [self.config.login, self.config.password, self.config.browser, self.config.TFA]
        
            driver = get_driver(self.config.browser) # Open the browser window
            login_to_quercus(driver, auth_info) # Log into quercus
            pra_id = get_assignment_id(driver, self.config.webpg_course, self.config.PRA_name) # Get the assignment ID
            print("Retrieved the assignment ID")
            time.sleep(2)
            
            # Change the name of the assignment to correctly input it later
            PRA_name = self.config.PRA_name + ' (' + pra_id + ')'
            out_path = os.path.join(os.path.dirname(__file__), self.config.file_out).replace('\\', '/')
        
            course_info = [self.config.webpg_course, self.config.PRA_name, out_path]
        
            self.write_to_csv(df, self.config.file_out) # Create the file with marks to export it later
            import_grades_to_quercus(driver, df, course_info) # Import grades to quercus
        elif file_extension == 'csv' and not self.config.quercus_grading:
            self.write_to_csv(df, self.config.file_out) # Create the file with marks to export it later
        else:
            raise ValueError(f"Unknown output file extension: {file_extension}")
    

#Main function

if __name__ == "__main__":
    config_file = "config.txt"
    removerows=0
    marking = Marking(config_file, removerows=removerows)
    marking.main()
